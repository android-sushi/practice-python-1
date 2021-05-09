import openpyxl as xl
import pprint
import csv

file_name = 'file_name'
data_set = []
data_set_child = []
header = ['A', 'B', 'C', 'D']

print('loading...')
wb = xl.load_workbook(file_name + '.xlsx')

for sheet in wb.sheetnames:
    data_no_cnt = 5
    data_flag_cnt = 4
    data_select_cnt = 8
    ws = wb[sheet]

    for row in ws.iter_rows():
        if data_no_cnt % 6 == 0:
            data_set_child.append(row[1].value)
        if data_flag_cnt % 6 == 0:
            data_set_child.append(row[1].value)
            data_set_child.append(row[0].value)
        if data_select_cnt % 6 == 0:
            data_set_child.append(row[0].value)
            data_set.append(data_set_child)
            data_set_child = []

        data_no_cnt += 1
        data_flag_cnt += 1
        data_select_cnt += 1

pprint.pprint(data_set)

with open(file_name + '.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(header)
    writer.writerows(data_set)
