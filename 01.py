import openpyxl as xl


def read_wb(make_sheet):
    file_pass = make_sheet + '.xlsx'
    read_data = xl.load_workbook(file_pass)
    return read_data


def make_combination_x(make_sheet):
    if make_sheet == 'A':
        return list(range(1, 7))
    elif make_sheet == 'B':
        return list(range(1, 8))
    elif make_sheet == 'C':
        return list(range(1, 9))
    elif make_sheet == 'D':
        return list(range(1, 10))


def make_combinations_y(ws):
    combinations_y = []
    columns_list = ('B', 'G', 'L', 'Q')
    row_list = (3, 15, 27, 39)
    for columns in columns_list:
        for row in row_list:
            combinations_y.append(ws[columns + str(row)].value)
    return combinations_y


def write_wb(active_make_sheet, row, lr_lists, make_sheet, combination, y_num, y_num_shift, num):
    cnt = 1

    for lr_list in lr_lists:
        active_make_sheet['A' + str(row)].value = str(make_sheet)
        active_make_sheet['B' + str(row)].value = combination
        active_make_sheet['C' + str(row)].value = y_num
        active_make_sheet['D' + str(row)].value = y_num_shift
        active_make_sheet['E' + str(row)].value = lr_list[0]
        active_make_sheet['F' + str(row)].value = lr_list[1]
        active_make_sheet['G' + str(row)].value = cnt

        active_make_sheet['H' + str(row)].value = '=IF(I' + str(row) + '<50,3,5)'
        active_make_sheet['I' + str(row)].value = num[y_num][num - 1]
        active_make_sheet['J' + str(row)].value = '="FOO"&CHAR(10)&A' + str(row) + '&B' + str(row) +\
                                                  '&C' + str(row) + '&E' + str(row) + '&F' + str(row)
        active_make_sheet['K' + str(row)].value = '="BAR"&CHAR(10)&A' + str(row) + '&B' + str(row) +\
                                                  '&C' + str(row) + '&G' + str(row)
        active_make_sheet['L' + str(row)].value = '="BAZ"&H' + str(row) + '&"QUX"'
        active_make_sheet['M' + str(row)].value = '=J' + str(row) + '&CHAR(10)&K' + str(row) + '&CHAR(10)&L' + str(row)

        cnt += 1
        row += 1

    return row


def main():
    wb = xl.load_workbook('read_wb.xlsx')

    pattern_11 = (
        (1, 'A'), (1, 'B'), (2, 'A'), (2, 'B'), (3, 'B'),
        (4, 'A'), (4, 'B'), (5, 'A'), (5, 'B'), (6, 'A'), (6, 'B')
    )
    pattern_12 = (
        (1, 'A'), (1, 'B'), (2, 'A'), (2, 'B'), (3, 'A'), (3, 'B'),
        (4, 'A'), (4, 'B'), (5, 'A'), (5, 'B'), (6, 'A'), (6, 'B')
    )

    xl_row = (list(range(3, 15)), list(range(15, 27)), list(range(27, 39)), list(range(39, 51)))
    xl_column = ('D', 'I', 'N', 'S')
    make_sheets = ('A', 'B', 'C', 'D')

    for make_sheet in make_sheets:
        read_data = read_wb(make_sheet)
        combinations = make_combination_x(make_sheet)
        active_make_sheet = wb[make_sheet]
        row = 2

        for combination_x in combinations:
            sub = make_sheet + str(combination_x)
            datasets = read_data[sub]
            combinations_y = make_combinations_y(datasets)
            num = {}
            n = 1

            for xl_column_num in xl_column:
                for xl_row_nums in xl_row:
                    num_list = []
                    for xl_row_num in xl_row_nums:
                        num_list.append(datasets[xl_column_num + str(xl_row_num)].value)
                    num[n] = num_list
                    n += 1

            for combination_y in enumerate(combinations_y):
                y_num = combination_y[0] + 1
                y_num_shift = combination_y[1]

                if combination_y[1] == 11:
                    row = write_wb(active_make_sheet, row, pattern_11, make_sheet,
                                   combination_x, y_num, y_num_shift, num)
                elif combination_y[1] == 12:
                    row = write_wb(active_make_sheet, row, pattern_12, make_sheet,
                                   combination_x, y_num, y_num_shift, num)

    wb.save('save_book.xlsx')


if __name__ == '__main__':
    main()
